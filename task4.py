from openpyxl import load_workbook
import math

# Load the workbook and target sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Initialize total and count
total_sum = 0
count = 0

# Loop through each row starting from row 2
for row in range(2, ws.max_row + 1):
    product = ws['E' + str(row)].value  # Column E = Produkts
    price = ws['F' + str(row)].value    # Column F = Cena

    # Check if product contains "LaserJet" and price is a number
    if isinstance(product, str) and "LaserJet" in product and isinstance(price, (int, float)):
        total_sum += price
        count += 1

# Calculate average and round down using math.floor
average_price = math.floor(total_sum / count) if count > 0 else 0

print("Average price (rounded down):", average_price)


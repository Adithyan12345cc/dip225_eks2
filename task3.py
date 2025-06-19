from openpyxl import load_workbook

# Load workbook and sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Initialize counter
count = 0

# Get total number of rows
max_row = ws.max_row

# Loop through all rows (skipping header)
for row in range(2, max_row + 1):
    address = ws['A' + str(row)].value  # Column A = Adrese
    city = ws['D' + str(row)].value     # Column D = Pilsēta

    # Check if conditions are met
    if isinstance(address, str) and "Adulienas iela" in address and city in ["Valmiera", "Saulkrasti"]:
        count += 1

print("Number of matching entries:", count)


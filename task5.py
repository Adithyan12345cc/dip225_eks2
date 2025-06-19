from openpyxl import load_workbook
import math

# Load the workbook and target sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Initialize total sum
total_sum = 0

# Loop through all rows starting from 2 (skip header)
for row in range(2, ws.max_row + 1):
    client = ws['G' + str(row)].value   # Column G = Klients
    quantity = ws['H' + str(row)].value # Column H = Skaits
    total = ws['I' + str(row)].value    # Column I = Kopā

    # Check if client is "Korporatīvais", quantity in range 40–50, and total is valid
    if client == "Korporatīvais" and isinstance(quantity, (int, float)) and 40 <= quantity <= 50 and isinstance(total, (int, float)):
        total_sum += total

# Round down the result
total_sum_rounded = math.floor(total_sum)

print("Total sum for 'Korporatīvais' clients with quantity 40–50 (rounded down):", total_sum_rounded)


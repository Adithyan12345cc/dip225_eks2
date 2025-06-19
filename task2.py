from openpyxl import load_workbook

# Load the workbook and the sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Initialize counter
count = 0

# Get the total number of rows
max_row = ws.max_row

# Loop through the rows, starting from 2 to skip the header
for row in range(2, max_row + 1):
    priority = ws['B' + str(row)].value  # Column B = Prioritāte
    delivery_date = ws['C' + str(row)].value  # Column C = Piegādes datums

    # Check if priority is "High" and delivery date is in 2015
    if priority == "High" and hasattr(delivery_date, 'year') and delivery_date.year == 2015:
        count += 1

print("Number of entries with High priority and delivery date in 2015:", count)


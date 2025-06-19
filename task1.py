from openpyxl import load_workbook

# Load the Excel file and the sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_1']

# Initialize counter
count = 0

# Get the total number of rows
max_row = ws.max_row

# Loop through rows (starting from row 2 to skip headers)
for row in range(2, max_row + 1):
    address = ws['A' + str(row)].value  # Column A = Address
    value = ws['B' + str(row)].value    # Column B = Skaits (Count)

    # Check if address starts with "Ain" and value is less than 40
    if isinstance(address, str) and address.startswith('Ain') and isinstance(value, (int, float)) and value < 40:
        count += 1

print("Number of matching entries:", count)

